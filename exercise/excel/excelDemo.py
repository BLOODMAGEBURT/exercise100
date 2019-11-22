# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     excelDemo
   Description :
   Author :       Administrator
   date：          2019/11/22 0022
-------------------------------------------------
   Change Activity:
                   2019/11/22 0022:
-------------------------------------------------
"""
import os

from openpyxl import load_workbook, Workbook


def save_cell_value(file_path):
    """
    存数据
    :param file_path:
    :return:
    """
    file_name_list = os.listdir(file_path)

    print(file_name_list)

    wb = Workbook()
    ws = wb.active
    ws.title = '智慧供应链'

    for file in file_name_list:
        name = get_cell_value(base_path + "\\" + file, 2, 1)
        value = get_cell_value(base_path + "\\" + file, 2, 2)
        row = [name, value]
        ws.append(row)

    wb.save('智慧供应链.xlsx')


def get_cell_value(path, row, col):
    """
    读取Excel的数据并打印出来
    :param path:
    :param row:
    :param col:
    :return:
    """
    wb = load_workbook(path)
    st = wb.active
    return st.cell(row, col).value


if __name__ == '__main__':
    base_path = "E:\项目管理2\电信平台2019\验收文档\最终验收\数据词典\数据词典\智慧供应链"
    save_cell_value(base_path)
    # get_cell_value(path,1, 2)
